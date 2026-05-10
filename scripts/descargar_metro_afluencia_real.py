"""
descargar_metro_afluencia_real.py — Descarga la afluencia real del Metro de
Medellín desde el portal de Datos Abiertos (ArcGIS Hub) y la convierte a un
CSV en formato largo, listo para Bronze.

Por qué este script:
    El portal entrega xlsx anuales en formato ancho (filas = día×línea, columnas =
    horas 4:00..23:00). Los IDs de los items rotan cuando el portal se actualiza,
    así que primero consultamos el feed DCAT (que siempre está fresco) y de ahí
    extraemos el itemID actual de cada año. Luego /sharing/rest/content/items/<id>/data
    devuelve el xlsx (con un User-Agent de navegador, si no responde 403).

Salida:
    data/raw/metro_afluencia/afluencia_metro_<YYYY>.csv
    Columnas (formato largo): fecha, linea, hora, pasajeros
    El esquema reemplaza al sintético previo (que mentía con "estación").
    Decisión documentada: el dato público es por línea, no por estación.

Uso (desde la raíz del repo, en host con Python 3.10+):
    pip install requests openpyxl
    python scripts/descargar_metro_afluencia_real.py
    python scripts/descargar_metro_afluencia_real.py --anios 2022 2023 2024
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from datetime import date, datetime, time
from pathlib import Path

try:
    import openpyxl
    import requests
except ImportError as exc:
    print(f"ERROR: falta dependencia ({exc.name}). Instalar:", file=sys.stderr)
    print("  pip install requests openpyxl", file=sys.stderr)
    sys.exit(1)

DCAT_URL = (
    "https://datosabiertos-metrodemedellin.opendata.arcgis.com/api/feed/dcat-us/1.1.json"
)
ITEM_DATA_URL = "https://www.arcgis.com/sharing/rest/content/items/{id}/data"
DESTINO_DIR = Path("data/raw/metro_afluencia")
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)
SESION = requests.Session()
SESION.headers.update({"User-Agent": USER_AGENT})


def _resolver_items_afluencia() -> dict[int, str]:
    """Año → itemID actual, leído del feed DCAT del portal."""
    resp = SESION.get(DCAT_URL, timeout=30)
    resp.raise_for_status()
    catalogo = resp.json()
    items: dict[int, str] = {}
    patron_anio = re.compile(r"Afluencia\s+Metro\s+(\d{4})", re.IGNORECASE)
    for ds in catalogo.get("dataset", []):
        m = patron_anio.match(ds.get("title", "").strip())
        if not m:
            continue
        anio = int(m.group(1))
        identifier = ds.get("identifier", "")
        m2 = re.search(r"id=([0-9a-f]{32})", identifier)
        if m2:
            items[anio] = m2.group(1)
    return items


def _descargar_xlsx(item_id: str, destino: Path) -> bool:
    if destino.exists() and destino.stat().st_size > 10_000:
        print(f"  ↺ ya existe: {destino.name}")
        return True
    url = ITEM_DATA_URL.format(id=item_id)
    print(f"  ⇣ {url} → {destino.name}")
    with SESION.get(url, stream=True, timeout=60) as r:
        r.raise_for_status()
        with destino.open("wb") as f:
            for chunk in r.iter_content(chunk_size=64 * 1024):
                if chunk:
                    f.write(chunk)
    return destino.stat().st_size > 0


def _normalizar_fecha(valor) -> date | None:
    """El xlsx trae fechas como datetime, str 'dd.mm.yyyy' o int 'aaaammdd'."""
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        s = valor.strip()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _xlsx_a_csv_largo(xlsx_path: Path, csv_path: Path) -> int:
    """Convierte un xlsx Metro de formato ancho (día×línea×hora) a CSV largo.

    Encabezados xlsx:
      fila 1: 'Día', 'Línea de Servicio', 'Hora de operación', None, ..., None
      fila 2: None, None, time(4,0), time(5,0), ..., time(23,0), 'Total general'
      filas 3+: '01.01.2024', 'LÍNEA 1', 180, 858, ..., 22785
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet = wb[wb.sheetnames[0]]
        filas = sheet.iter_rows(values_only=True)

        next(filas)            # encabezado
        fila_horas = next(filas)
        horas: list[int] = []
        for celda in fila_horas[2:]:
            if isinstance(celda, time):
                horas.append(celda.hour)
            else:
                horas.append(-1)  # columna 'Total general' u otros

        escritas = 0
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["fecha", "linea", "hora", "pasajeros"])
            for row in filas:
                if not row or row[0] is None:
                    continue
                fecha = _normalizar_fecha(row[0])
                linea = (row[1] or "").strip()
                if fecha is None or not linea:
                    continue
                for col_idx, hora in enumerate(horas):
                    if hora == -1:
                        continue
                    valor = row[2 + col_idx]
                    if valor is None:
                        continue
                    try:
                        pax = int(valor)
                    except (TypeError, ValueError):
                        continue
                    if pax < 0:
                        continue
                    w.writerow([fecha.isoformat(), linea, hora, pax])
                    escritas += 1
        return escritas
    finally:
        wb.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--anios", type=int, nargs="+", default=[2022, 2023, 2024])
    parser.add_argument(
        "--mantener-xlsx",
        action="store_true",
        help="No borrar los xlsx descargados tras convertirlos a CSV",
    )
    args = parser.parse_args()

    DESTINO_DIR.mkdir(parents=True, exist_ok=True)

    print("→ Resolviendo item IDs actuales desde DCAT…")
    catalogo = _resolver_items_afluencia()
    if not catalogo:
        print("ERROR: el feed DCAT no listó datasets de afluencia.", file=sys.stderr)
        return 2
    print(f"  Años publicados: {sorted(catalogo)}")

    fallidos: list[int] = []
    for anio in args.anios:
        print(f"\n=== Afluencia {anio} ===")
        if anio not in catalogo:
            print(f"  ⚠ no publicado en el portal — se omite")
            fallidos.append(anio)
            continue
        item_id = catalogo[anio]
        xlsx = DESTINO_DIR / f"afluencia_metro_{anio}.xlsx"
        try:
            _descargar_xlsx(item_id, xlsx)
        except Exception as exc:
            print(f"  ✗ descarga falló: {exc}")
            fallidos.append(anio)
            continue
        csv_out = DESTINO_DIR / f"afluencia_metro_{anio}.csv"
        try:
            n = _xlsx_a_csv_largo(xlsx, csv_out)
            print(f"  ✓ {csv_out.name}  ({n:,} filas largas)")
        except Exception as exc:
            print(f"  ✗ conversión falló: {exc}")
            fallidos.append(anio)
            continue
        if not args.mantener_xlsx:
            xlsx.unlink()

    if fallidos:
        print(f"\n⚠ Años con fallos: {fallidos}", file=sys.stderr)
        return 1
    print("\n✅ Metro afluencia real lista en data/raw/metro_afluencia/")
    return 0


if __name__ == "__main__":
    sys.exit(main())
